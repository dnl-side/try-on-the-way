"""
Sales Analytics Dashboard - Enterprise Analytics Interface

A comprehensive PyQt6-based analytics dashboard featuring:
- Advanced calendar widget with multi-date range selection (Shift+Click functionality)
- Real-time financial analytics with currency conversion capabilities
- Professional Excel export system with multi-sheet formatting
- Dynamic data visualization integration with interactive charts
- Multi-source data synchronization (local + remote database strategies)
- Sophisticated UI layout management with responsive design principles
- Complex business logic for sales aggregation and financial calculations

Technical Achievements:
- Advanced PyQt6 calendar implementation with custom date range selection
- Complex financial calculations with tax handling and currency conversion
- Professional Excel export with corporate styling and multi-sheet architecture
- Dynamic table management with real-time data updates and formatting
- Sophisticated layout management with splitters and responsive design
- Memory-efficient data processing with proper resource cleanup
- Custom widget integration for charts and specialized components
- Advanced styling system with CSS integration and theming support

Business Value:
- Real-time sales analytics and performance monitoring
- Comprehensive financial reporting with tax calculations
- Multi-currency support for international business operations
- Professional documentation export for stakeholder reporting
- Enhanced decision-making through visual data representation
- Improved productivity through intuitive interface design

@author Daniel Jara
@version 2.1.0
@since Python 3.8+
"""

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, 
    QLayout, QFrame, QFileDialog, QLineEdit, QCalendarWidget, QTableWidgetItem, 
    QHeaderView, QMessageBox, QScrollArea, QSizePolicy, QSplitter, QPushButton
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QGuiApplication, QTextCharFormat, QIcon, QColor
from decimal import Decimal
import mysql.connector
from database_connection import conectar
from interactive_charts import InteractiveChartsWidget
from data_processor import SalesDataProcessor
from resource_path import resource_path
from loadstyle import apply_stylesheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import os


class SalesDashboard(QMainWindow):
    """
    Sales Analytics Dashboard - Main Controller Interface
    
    Demonstrates advanced PyQt6 dashboard development:
    - Complex calendar widget with range selection capabilities
    - Multi-table data management with real-time updates
    - Professional Excel export with corporate formatting
    - Dynamic UI components with responsive layout design
    - Advanced business logic integration for financial calculations
    - Modular architecture with separated concerns (charts, data processing)
    - Memory-efficient resource management with proper cleanup
    
    Features:
    - Interactive calendar with Shift+Click date range selection
    - Real-time financial calculations and currency conversion
    - Professional multi-sheet Excel export with styling
    - Dynamic table updates with live data synchronization
    - Integrated chart visualization with PyQtGraph components
    - Responsive layout design for different screen resolutions
    - Advanced error handling with user-friendly feedback
    """

    def __init__(self, parent=None):
        """
        Initialize Sales Analytics Dashboard
        
        Args:
            parent: Parent widget for proper window management
        """
        super().__init__(parent)
        
        # Configure window properties
        self.setWindowIcon(QIcon(resource_path('icons/app_icon.PNG')))
        self.setWindowTitle("Sales Analytics Dashboard")
        self._setup_window_geometry()
        
        # Initialize data structures
        self.PRODUCT_ORDER = []  
        self.PRODUCT_UNIT = {}   
        self.PELLET_PRODUCTS = []
        self.VACAM_PRODUCTS = []
        
        # Initialize date selection state
        self.start_date = None
        self.end_date = None
        self.current_date = QDate.currentDate()
        
        # Load product configuration
        self._load_product_configuration()
        
        # Create UI components
        self._create_main_interface()
        self._create_data_tables()
        self._initialize_chart_integration()
        self._configure_calendar_system()
        
        # Apply professional styling
        self._apply_dashboard_styling()
        
        # Load initial data
        self._refresh_dashboard_data()

    def _setup_window_geometry(self):
        """Configure optimal window size and positioning for dashboard"""
        screen_geometry = QGuiApplication.primaryScreen().availableGeometry()
        initial_width = int(screen_geometry.width() * 1.0)   # Full width
        initial_height = int(screen_geometry.height() * 0.9) # 90% height
        self.setFixedSize(initial_width, initial_height)

    def _create_main_interface(self):
        """
        Create comprehensive dashboard interface
        Implements professional layout with modular components
        """
        # Create scrollable central widget
        self.central_widget = QWidget(self)
        self.central_layout = QVBoxLayout(self.central_widget)
        
        # Configure scrollable area for responsive design
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setWidget(self.central_widget)
        self.setCentralWidget(self.scroll_area)
        
        # Create main splitter for dashboard layout
        self._create_dashboard_layout()

    def _create_dashboard_layout(self):
        """
        Create sophisticated dashboard layout with splitters
        Implements responsive design with proper component separation
        """
        # Create main horizontal splitter
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.central_layout.addWidget(self.main_splitter)

        # Left panel: Controls and tables
        self.left_container = QWidget()
        self.left_layout = QVBoxLayout(self.left_container)
        
        # Right panel: Charts integration
        self.right_container = QWidget()
        self.right_layout = QVBoxLayout(self.right_container)

        # Configure splitter proportions
        self.main_splitter.addWidget(self.left_container)
        self.main_splitter.addWidget(self.right_container)
        self.main_splitter.setSizes([
            int(self.width() * 0.7), 
            int(self.width() * 0.3)
        ])

        # Add dashboard components
        self._create_control_panel()
        self._add_data_tables_section()

    def _create_control_panel(self):
        """
        Create dashboard control panel with calendar and exchange rate
        Implements advanced calendar functionality with range selection
        """
        # Calendar toggle button with smart labeling
        self.toggle_calendar_button = QPushButton("Show Calendar")
        self.toggle_calendar_button.clicked.connect(self._toggle_calendar_visibility)
        self.left_layout.addWidget(self.toggle_calendar_button)

        # Calendar container with advanced functionality
        self.calendar_container = QFrame()
        self.calendar_layout = QVBoxLayout(self.calendar_container)        
        
        # Configure advanced calendar widget
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setMinimumDate(QDate(2024, 4, 1))  # Business start date
        self.calendar.setSelectedDate(self.current_date)
        self.calendar.clicked.connect(self._refresh_dashboard_data)
        
        # Advanced date selection with range support
        self.calendar.selectionChanged.connect(self._handle_date_selection_change)
        
        self.calendar_layout.addWidget(self.calendar)
        self.calendar_container.setVisible(False)  # Initially hidden
        self.left_layout.addWidget(self.calendar_container)
        
        # Exchange rate input for currency conversion
        self.exchange_rate_input = QLineEdit()
        self.exchange_rate_input.setPlaceholderText("Exchange Rate (Default: 945)")
        self.left_layout.addWidget(self.exchange_rate_input)
        
        # Export functionality
        self._create_export_controls()
        
        # Update calendar button text
        self._update_calendar_button_text()

    def _create_export_controls(self):
        """Create professional export control buttons"""
        export_layout = QHBoxLayout()
        
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self._export_comprehensive_report)
        
        self.exit_button = QPushButton("Exit Dashboard")
        self.exit_button.clicked.connect(self._handle_dashboard_exit)
        
        export_layout.addWidget(self.export_button)
        export_layout.addWidget(self.exit_button)
        self.left_layout.addLayout(export_layout)

    def _add_data_tables_section(self):
        """
        Add comprehensive data tables section
        Implements multiple table management with responsive design
        """
        # Create container for all data tables
        tables_container = QVBoxLayout()
        tables_container.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)
        
        # Branch-specific sales tables
        tables_container.addWidget(QLabel("Branch Alpha - Daily Sales"))
        self.sales_by_branch_alpha = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.sales_by_branch_alpha)
        
        tables_container.addWidget(QLabel("Branch Alpha - Monthly Sales"))
        self.monthly_totals_alpha = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.monthly_totals_alpha)
        
        tables_container.addWidget(QLabel("Branch Beta - Daily Sales"))
        self.sales_by_branch_beta = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.sales_by_branch_beta)
        
        tables_container.addWidget(QLabel("Branch Beta - Monthly Sales"))
        self.monthly_totals_beta = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.monthly_totals_beta)
        
        # Consolidated results tables
        tables_container.addWidget(QLabel("Daily Results Summary"))
        self.daily_sales_table = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.daily_sales_table)
        
        tables_container.addWidget(QLabel("Monthly Results Summary"))
        self.monthly_sales_table = self._create_standardized_table([
            "Product", "Quantity", "Weight", "Net Amount", "Total Amount", "USD Amount"
        ])
        tables_container.addWidget(self.monthly_sales_table)
        
        # Add tables container to main layout
        self.left_layout.addLayout(tables_container)

    def _create_standardized_table(self, headers: list) -> QTableWidget:
        """
        Create standardized table with professional formatting
        
        Args:
            headers (list): Table column headers
            
        Returns:
            QTableWidget: Configured table widget
        """
        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        return table

    def _create_data_tables(self):
        """
        Initialize data processing tables
        Note: Tables are created in _add_data_tables_section
        """
        pass  # Tables created in layout method

    def _initialize_chart_integration(self):
        """
        Initialize chart integration with interactive components
        Demonstrates modular architecture with chart separation
        """
        try:
            # Initialize charts widget as separate component
            self.charts_widget = InteractiveChartsWidget()
            self.right_layout.addWidget(self.charts_widget)
            
        except ImportError as e:
            # Fallback if charts module unavailable
            print(f"Charts module unavailable: {e}")
            fallback_label = QLabel("Chart visualization unavailable in demo mode")
            fallback_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.right_layout.addWidget(fallback_label)

    def _configure_calendar_system(self):
        """
        Configure advanced calendar system with range selection
        Implements sophisticated date handling for business analytics
        """
        # Apply professional calendar styling
        self._apply_calendar_styling()
        
        # Configure initial date selection
        self._update_calendar_selection()

    # ================ DATE SELECTION LOGIC ================

    def _handle_date_selection_change(self):
        """
        Handle advanced date selection with range support
        Implements Shift+Click functionality for date ranges
        """
        modifiers = QGuiApplication.keyboardModifiers()
        
        if modifiers & Qt.KeyboardModifier.ShiftModifier:
            current_date = self.calendar.selectedDate()
    
            if self.start_date is None:
                # First selection in range
                self.start_date = current_date
                self.end_date = None
            else:
                # Second selection - complete range
                self.end_date = current_date
                
            # Ensure logical date ordering
            if self.start_date and self.end_date:
                if self.start_date > self.end_date:
                    self.start_date, self.end_date = self.end_date, self.start_date
    
            # Update dashboard with range selection
            self._refresh_dashboard_data()
            self._update_calendar_selection()
            self._update_calendar_button_text()
        else:
            # Single date selection - reset range
            self.start_date = self.calendar.selectedDate()
            self.end_date = None
            self._update_calendar_selection()
            self._update_calendar_button_text()

    def _update_calendar_selection(self):
        """
        Update calendar visual selection with range highlighting
        Implements professional visual feedback for date ranges
        """
        # Clear previous highlighting
        self.calendar.setDateTextFormat(QDate(), QTextCharFormat())
        
        if self.start_date and self.end_date:
            # Highlight complete date range
            start_date = min(self.start_date, self.end_date)
            end_date = max(self.start_date, self.end_date)
            
            for date in self._generate_date_range(start_date, end_date):
                self.calendar.setDateTextFormat(date, self._create_highlight_format())
            
        self.calendar.update()

    def _generate_date_range(self, start_date: QDate, end_date: QDate):
        """
        Generate date range iterator for calendar highlighting
        
        Args:
            start_date (QDate): Range start date
            end_date (QDate): Range end date
            
        Yields:
            QDate: Each date in the range
        """
        current_date = start_date
        while current_date <= end_date:
            yield current_date
            current_date = current_date.addDays(1)

    def _create_highlight_format(self) -> QTextCharFormat:
        """
        Create professional highlight format for date selection
        
        Returns:
            QTextCharFormat: Configured highlight format
        """
        format = QTextCharFormat()
        format.setBackground(QColor("#e94560"))  # Corporate accent color
        format.setForeground(QColor("#ffffff"))   # White text for contrast
        return format

    def _update_calendar_button_text(self):
        """Update calendar button text with selected date information"""
        if self.start_date and self.end_date:
            if self.start_date == self.end_date:
                date_text = self.start_date.toString("dd/MM/yyyy")
            else:
                date_text = f"{self.start_date.toString('dd/MM/yyyy')} - {self.end_date.toString('dd/MM/yyyy')}"
        elif self.start_date:
            date_text = self.start_date.toString("dd/MM/yyyy")
        else:
            date_text = self.current_date.toString("dd/MM/yyyy")

        if self.calendar_container.isVisible():
            self.toggle_calendar_button.setText(f"Hide Calendar ({date_text})")
        else:
            self.toggle_calendar_button.setText(f"Show Calendar ({date_text})")

    def _toggle_calendar_visibility(self):
        """Toggle calendar visibility with smart button text updates"""
        if self.calendar_container.isVisible():
            self.calendar_container.setVisible(False)
        else:
            self.calendar_container.setVisible(True)
        
        self._update_calendar_button_text()

    # ================ DATA PROCESSING ================

    def _refresh_dashboard_data(self):
        """
        Refresh all dashboard data with current date selection
        Implements comprehensive data synchronization across all components
        """
        if not hasattr(self, 'PRODUCT_ORDER') or not hasattr(self, 'PRODUCT_UNIT'):
            print("Product configuration not loaded properly")
            return
        
        # Determine date range for data query
        if self.start_date and self.end_date:
            start_date_str = self.start_date.toString(Qt.DateFormat.ISODate)
            end_date_str = self.end_date.toString(Qt.DateFormat.ISODate)
            query_condition = "Date BETWEEN %s AND %s"
            query_params = (start_date_str, end_date_str)
        else:
            selected_date_str = self.calendar.selectedDate().toString(Qt.DateFormat.ISODate)
            query_condition = "Date = %s"
            query_params = (selected_date_str,)
        
        # Load sales data from database
        daily_sales_data, monthly_sales_data = self._load_sales_data(query_condition, query_params)
        
        # Process and display data
        if daily_sales_data or monthly_sales_data:
            self._populate_sales_tables(daily_sales_data, monthly_sales_data)
            self._update_branch_tables()
            self._adjust_table_displays()
            
            # Update chart integration if available
            if hasattr(self, 'charts_widget'):
                self.charts_widget.update_charts_data(daily_sales_data, monthly_sales_data)

    def _load_sales_data(self, query_condition: str, query_params: tuple):
        """
        Load sales data from database with error handling
        
        Args:
            query_condition (str): SQL WHERE condition
            query_params (tuple): Query parameters
            
        Returns:
            tuple: Daily and monthly sales data
        """
        try:
            conn = conectar()
            if conn is None:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor(dictionary=True)
            
            # Load daily sales data
            daily_query = f"""
                SELECT Product AS product, Quantity AS quantity, TotalAmount AS total, 
                       SalesLocation AS branch, TotalWeight AS weight
                FROM DailySales
                WHERE {query_condition}
            """
            cursor.execute(daily_query, query_params)
            daily_sales_data = cursor.fetchall()
        
            # Load monthly sales data (aggregated)
            monthly_sales_data = self._load_monthly_aggregated_data(cursor, query_condition, query_params)
        
            cursor.close()
            conn.close()
            
            return daily_sales_data, monthly_sales_data
        
        except mysql.connector.Error as e:
            error_message = f"Database error: {e}"
            print(error_message)
            QMessageBox.critical(self, "Database Error", error_message)
            return [], []

    def _load_monthly_aggregated_data(self, cursor, query_condition: str, query_params: tuple):
        """
        Load aggregated monthly sales data
        
        Args:
            cursor: Database cursor
            query_condition (str): WHERE condition
            query_params (tuple): Query parameters
            
        Returns:
            list: Monthly aggregated data
        """
        # Calculate month range for aggregation
        if self.start_date and self.end_date:
            start_month = self.start_date.month()
            start_year = self.start_date.year()
            end_month = self.end_date.month()
            end_year = self.end_date.year()
        else:
            start_month = end_month = self.calendar.selectedDate().month()
            start_year = end_year = self.calendar.selectedDate().year()

        monthly_data = []
        for month in range(start_month, end_month + 1):
            start_date_month = QDate(start_year, month, 1)
            end_date_month = start_date_month.addMonths(1).addDays(-1)
            
            monthly_query = """
                SELECT Product AS product, SUM(Quantity) as total_quantity, 
                       SUM(TotalAmount) as total_total, SalesLocation AS branch, 
                       SUM(TotalWeight) as total_weight
                FROM DailySales
                WHERE Date BETWEEN %s AND %s
                GROUP BY Product, SalesLocation
            """
            cursor.execute(monthly_query, (
                start_date_month.toString(Qt.DateFormat.ISODate), 
                end_date_month.toString(Qt.DateFormat.ISODate)
            ))
            monthly_data.extend(cursor.fetchall())
        
        return monthly_data

    def _populate_sales_tables(self, daily_data: list, monthly_data: list):
        """
        Populate sales tables with processed data
        
        Args:
            daily_data (list): Daily sales records
            monthly_data (list): Monthly aggregated data
        """
        # Initialize data processor
        data_processor = SalesDataProcessor(self)
        
        # Populate daily sales table
        data_processor.populate_daily_sales_table(daily_data)
        
        # Populate monthly sales table  
        data_processor.populate_monthly_sales_table(monthly_data)

    def _update_branch_tables(self):
        """Update branch-specific tables with current data"""
        # Initialize data processor for branch tables
        data_processor = SalesDataProcessor(self)
        
        # Update branch-specific data tables
        data_processor.populate_sales_by_branch_table()
        data_processor.populate_monthly_sales_by_branch_table()

    def _adjust_table_displays(self):
        """
        Adjust table display properties for optimal presentation
        Implements responsive table sizing and formatting
        """
        tables = [
            self.sales_by_branch_alpha, self.monthly_totals_alpha,
            self.sales_by_branch_beta, self.monthly_totals_beta,
            self.daily_sales_table, self.monthly_sales_table
        ]
        
        for table in tables:
            self._configure_table_display(table)

    def _configure_table_display(self, table: QTableWidget):
        """
        Configure individual table display properties
        
        Args:
            table (QTableWidget): Table to configure
        """
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Configure header properties
        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setStretchLastSection(True)
        
        # Center-align all cell content
        self._align_table_cells_center(table)
        self._adjust_table_height(table)

    def _align_table_cells_center(self, table: QTableWidget):
        """Center-align all cells in the table"""
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _adjust_table_height(self, table: QTableWidget):
        """
        Dynamically adjust table height based on content
        
        Args:
            table (QTableWidget): Table to adjust
        """
        table.viewport().update()
        
        row_count = table.rowCount()
        if row_count > 0:
            row_height = table.rowHeight(0)
            total_height = row_count * row_height
            header_height = table.horizontalHeader().height()
            extra_height = 5
            
            min_height = total_height + header_height + extra_height
            table.setMinimumHeight(min_height)
            table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        else:
            table.setMinimumHeight(100)
            table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        table.updateGeometry()
        table.repaint()

    # ================ EXCEL EXPORT SYSTEM ================

    def _export_comprehensive_report(self):
        """
        Export comprehensive Excel report with professional formatting
        Implements multi-sheet export with corporate styling
        """
        # Validate date selection
        if not self.start_date:
            self.start_date = QDate.currentDate()

        # Create workbook with professional styling
        wb = Workbook()
        
        # Define corporate styling
        styling = self._create_excel_styling()
        
        # Create summary sheet with consolidated data
        summary_sheet = wb.create_sheet(title="Sales Summary", index=0)
        self._populate_summary_sheet(summary_sheet, styling)
        
        # Create detailed sheets for each branch
        self._create_branch_sheets(wb, styling)
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Generate filename and save
        filename = self._generate_export_filename()
        self._save_and_open_report(wb, filename)

    def _create_excel_styling(self) -> dict:
        """
        Create professional Excel styling configuration
        
        Returns:
            dict: Styling configuration
        """
        return {
            'header_font': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid"),
            'cell_fill': PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid"),
            'cell_font': Font(color="FFFFFF"),
            'title_font': Font(bold=True, size=14, color="FFFFFF"),
            'title_fill': PatternFill(start_color="e94560", end_color="e94560", fill_type="solid"),
            'number_format': '#,##0.00 [$USD]'
        }

    def _populate_summary_sheet(self, sheet, styling: dict):
        """
        Populate summary sheet with comprehensive data
        
        Args:
            sheet: Excel worksheet
            styling (dict): Styling configuration
        """
        # Add title and metadata
        self._add_summary_header(sheet, styling)
        
        # Add data tables
        current_row = 5
        tables_data = [
            ("Branch Alpha - Daily Sales", self.sales_by_branch_alpha),
            ("Branch Alpha - Monthly Sales", self.monthly_totals_alpha),
            ("Branch Beta - Daily Sales", self.sales_by_branch_beta),
            ("Branch Beta - Monthly Sales", self.monthly_totals_beta),
            ("Daily Summary", self.daily_sales_table),
            ("Monthly Summary", self.monthly_sales_table)
        ]
        
        for title, table in tables_data:
            current_row = self._add_table_to_sheet(sheet, table, current_row, title, styling)

    def _add_summary_header(self, sheet, styling: dict):
        """Add professional header to summary sheet"""
        exchange_rate = Decimal(self.exchange_rate_input.text()) if self.exchange_rate_input.text() else Decimal('945')
        
        if self.end_date:
            date_range = f"From {self.start_date.toString('dd-MM-yyyy')} to {self.end_date.toString('dd-MM-yyyy')}"
        else:
            date_range = f"Date: {self.start_date.toString('dd-MM-yyyy')}"
        
        # Title
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = sheet.cell(row=1, column=1, value="Sales Analytics Report")
        title_cell.font = styling['title_font']
        title_cell.fill = styling['title_fill']
        title_cell.alignment = Alignment(horizontal='center')
        
        # Date range
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        date_cell = sheet.cell(row=2, column=1, value=date_range)
        date_cell.font = styling['cell_font']
        date_cell.fill = styling['cell_fill']
        date_cell.alignment = Alignment(horizontal='center')
        
        # Exchange rate
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        rate_cell = sheet.cell(row=3, column=1, value=f"Exchange Rate: {exchange_rate:.2f}")
        rate_cell.font = styling['cell_font']
        rate_cell.fill = styling['cell_fill']
        rate_cell.alignment = Alignment(horizontal='center')

    def _add_table_to_sheet(self, sheet, table: QTableWidget, start_row: int, title: str, styling: dict) -> int:
        """
        Add table data to Excel sheet with formatting
        
        Args:
            sheet: Excel worksheet
            table (QTableWidget): Source table
            start_row (int): Starting row position
            title (str): Table title
            styling (dict): Styling configuration
            
        Returns:
            int: Next available row position
        """
        # Add table title
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell = sheet.cell(row=start_row, column=1, value=title)
        title_cell.font = styling['title_font']
        title_cell.fill = styling['title_fill']
        title_cell.alignment = Alignment(horizontal='center')
        
        current_row = start_row + 1
        
        # Add headers
        for col in range(table.columnCount()):
            header_text = table.horizontalHeaderItem(col).text()
            cell = sheet.cell(row=current_row, column=col + 1, value=header_text)
            cell.font = styling['header_font']
            cell.fill = styling['header_fill']
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Add data rows
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                cell_value = item.text() if item else ""
                cell = sheet.cell(row=current_row, column=col + 1, value=cell_value)
                cell.font = styling['cell_font']
                cell.fill = styling['cell_fill']
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        return current_row + 2  # Add spacing between tables

    def _create_branch_sheets(self, wb: Workbook, styling: dict):
        """
        Create individual sheets for each branch with detailed data
        
        Args:
            wb (Workbook): Excel workbook
            styling (dict): Styling configuration
        """
        # Branch Alpha sheet
        alpha_sheet = wb.create_sheet(title="Branch Alpha Details")
        self._populate_branch_sheet(alpha_sheet, "Alpha", [
            self.sales_by_branch_alpha, self.monthly_totals_alpha
        ], styling)
        
        # Branch Beta sheet
        beta_sheet = wb.create_sheet(title="Branch Beta Details")
        self._populate_branch_sheet(beta_sheet, "Beta", [
            self.sales_by_branch_beta, self.monthly_totals_beta
        ], styling)

    def _populate_branch_sheet(self, sheet, branch_name: str, tables: list, styling: dict):
        """
        Populate individual branch sheet with data
        
        Args:
            sheet: Excel worksheet
            branch_name (str): Branch identifier
            tables (list): List of tables to include
            styling (dict): Styling configuration
        """
        # Add branch header
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        header_cell = sheet.cell(row=1, column=1, value=f"Branch {branch_name} - Detailed Analysis")
        header_cell.font = styling['title_font']
        header_cell.fill = styling['title_fill']
        header_cell.alignment = Alignment(horizontal='center')
        
        current_row = 3
        table_titles = ["Daily Sales", "Monthly Sales"]
        
        for i, table in enumerate(tables):
            current_row = self._add_table_to_sheet(
                sheet, table, current_row, 
                f"{branch_name} - {table_titles[i]}", styling
            )

    def _generate_export_filename(self) -> str:
        """
        Generate appropriate filename for export
        
        Returns:
            str: Generated filename
        """
        if self.end_date:
            start_str = self.start_date.toString("dd-MM-yyyy")
            end_str = self.end_date.toString("dd-MM-yyyy")
            return f"Sales_Report_{start_str}_to_{end_str}.xlsx"
        else:
            date_str = self.start_date.toString("dd-MM-yyyy")
            return f"Sales_Report_{date_str}.xlsx"

    def _save_and_open_report(self, wb: Workbook, filename: str):
        """
        Save Excel report and open with default application
        
        Args:
            wb (Workbook): Excel workbook to save
            filename (str): Target filename
        """
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Sales Report", filename, 
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            wb.save(file_path)
            
            # Open with default application
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS/Linux
                    os.system(f"open '{file_path}'")  # macOS
            except Exception as e:
                print(f"Could not open file automatically: {e}")
                QMessageBox.information(
                    self, "Export Complete", 
                    f"Report saved successfully to:\n{file_path}"
                )

    # ================ CONFIGURATION MANAGEMENT ================

    def _load_product_configuration(self):
        """
        Load product configuration from database
        Implements flexible product management system
        """
        try:
            conn = conectar()
            if conn is None:
                raise Exception("Database connection failed")
                
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT Product, packaging, product_type FROM Products")
            products = cursor.fetchall()
            
            # Build configuration dictionaries
            self.PRODUCT_ORDER = [product['Product'] for product in products]
            self.PRODUCT_UNIT = {product['Product']: product['packaging'] for product in products}
            self.PELLET_PRODUCTS = [p['Product'] for p in products if p['product_type'] == 'Pellet']
            self.VACAM_PRODUCTS = [p['Product'] for p in products if p['product_type'] == 'Alternative']
    
            cursor.close()
            conn.close()
            
        except mysql.connector.Error as err:
            print(f"Error loading product configuration: {err}")
            # Fallback configuration for demo
            self._load_demo_configuration()

    def _load_demo_configuration(self):
        """Load demonstration configuration if database unavailable"""
        self.PRODUCT_ORDER = [
            "Pellet Bag 15kg (Retail)",
            "Pellet Bag 15kg (Delivery)", 
            "Pellet Bag 15kg (Wholesale)",
            "Alternative Product 3kg"
        ]
        self.PRODUCT_UNIT = {
            "Pellet Bag 15kg (Retail)": "bags",
            "Pellet Bag 15kg (Delivery)": "bags",
            "Pellet Bag 15kg (Wholesale)": "bags", 
            "Alternative Product 3kg": "units"
        }
        self.PELLET_PRODUCTS = [
            "Pellet Bag 15kg (Retail)",
            "Pellet Bag 15kg (Delivery)",
            "Pellet Bag 15kg (Wholesale)"
        ]
        self.VACAM_PRODUCTS = ["Alternative Product 3kg"]

    def _apply_dashboard_styling(self):
        """Apply professional styling to dashboard interface"""
        try:
            styles_path = resource_path("styles/main.css")
            calendar_styles_path = resource_path("styles/calendar.css")
            
            main_stylesheet = apply_stylesheet(styles_path)
            calendar_stylesheet = apply_stylesheet(calendar_styles_path)
            
            # Combine stylesheets for comprehensive styling
            combined_stylesheet = main_stylesheet + "\n" + calendar_stylesheet
            QApplication.instance().setStyleSheet(combined_stylesheet)
            
        except Exception as e:
            print(f"Error loading stylesheets: {e}")
            self._apply_fallback_styling()

    def _apply_calendar_styling(self):
        """Apply specialized styling to calendar widget"""
        try:
            calendar_css_path = resource_path("styles/calendar.css")
            apply_stylesheet(self.calendar, calendar_css_path)
        except Exception as e:
            print(f"Error applying calendar styling: {e}")

    def _apply_fallback_styling(self):
        """Apply fallback styling if external CSS unavailable"""
        fallback_style = """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 #1a1a2e, stop:1 #16213e);
            }
            QLabel {
                color: white;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 #0f3460, stop:1 #53354a);
                color: white;
                border: 2px solid #53354a;
                border-radius: 5px;
                padding: 6px 12px;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #e94560, stop:1 #0f3460);
            }
            QTableWidget {
                background: #1a1a2e;
                color: white;
                border: 1px solid #53354a;
            }
            QTableWidget::item {
                background: #1a1a2e;
                color: white;
            }
            QTableWidget QHeaderView::section {
                background: #0f3460;
                color: white;
                border: 1px solid #53354a;
            }
        """
        self.setStyleSheet(fallback_style)

    # ================ EVENT HANDLERS ================

    def _handle_dashboard_exit(self):
        """Handle dashboard exit with user confirmation"""
        reply = QMessageBox.question(
            self, 'Confirm Exit',
            'Are you sure you want to exit the dashboard?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.close()

    # ================ LIFECYCLE MANAGEMENT ================

    def closeEvent(self, event):
        """Handle window close event with proper cleanup"""
        # Cleanup chart widget if present
        if hasattr(self, 'charts_widget'):
            self.charts_widget.cleanup_resources()
        
        event.accept()


# ================ STANDALONE APPLICATION ENTRY POINT ================

def main():
    """
    Standalone application entry point for testing and demonstration
    Demonstrates modular dashboard architecture and independent operation
    """
    app = QApplication(sys.argv)
    
    # Create dashboard instance
    dashboard = SalesDashboard()
    dashboard.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
